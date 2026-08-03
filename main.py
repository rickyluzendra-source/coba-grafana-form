from fastapi import FastAPI, Response
from fastapi.middleware.cors import CORSMiddleware
import openpyxl
from openpyxl.styles import Font
from io import BytesIO

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Penampung data sementara di memory server
database_penampung = []

@app.post("/submit-data")
async def submit_data(data: dict):
    """Menampung inputan dari Grafana"""
    database_penampung.append(data)
    return {
        "status": "success", 
        "message": f"Data berhasil disimpan! Total data: {len(database_penampung)}"
    }

@app.post("/download-all")
async def download_all(payload: dict = None):
    """Mendownload SELURUH data tersimpan menjadi 1 File Excel dengan format rapi"""
    wb = openpyxl.Workbook()
    
    # ---------------------------------------------------------
    # SHEET 1: DATA MASTER
    # ---------------------------------------------------------
    ws_master = wb.active
    ws_master.title = "Data Master"
    ws_master.append(["ID", "Unit Bisnis", "Kategori", "Volume Sales", "Harga/Unit", "Biaya Direct/Unit"])
    
    for idx, item in enumerate(database_penampung, start=1):
        ws_master.append([
            f"TRX-{idx:03d}",
            item.get("unit_bisnis", "Branch Utama"),
            item.get("kategori", "General"),
            float(item.get("volume", 0)),
            float(item.get("harga_unit", 0)),
            float(item.get("biaya_direct", 0))
        ])

    # ---------------------------------------------------------
    # SHEET 2: HASIL KALKULASI
    # ---------------------------------------------------------
    ws_calc = wb.create_sheet(title="Hasil Kalkulasi")
    ws_calc.append(["ID", "Unit Bisnis", "Gross Revenue (IDR)", "Total Direct Cost (IDR)", "Tax Overhead (IDR)", "Net Profit (IDR)", "Margin %"])
    
    tax_rate = 0.11
    total_rows = len(database_penampung)
    
    # Loop untuk memasukkan rumus dan formatting per baris
    for r in range(2, total_rows + 2):
        ws_calc.append([
            f"='Data Master'!A{r}",
            f"='Data Master'!B{r}",
            f"='Data Master'!D{r}*'Data Master'!E{r}",  # Gross Revenue
            f"='Data Master'!D{r}*'Data Master'!F{r}",  # Direct Cost
            f"=C{r}*{tax_rate}",                       # Tax (11%)
            f"=C{r}-D{r}-E{r}",                        # Net Profit
            f"=F{r}/C{r}"                               # Margin %
        ])
        
        # Format Angka per Baris
        ws_calc[f"C{r}"].number_format = '#,##0'
        ws_calc[f"D{r}"].number_format = '#,##0'
        ws_calc[f"E{r}"].number_format = '#,##0'
        ws_calc[f"F{r}"].number_format = '#,##0'
        ws_calc[f"G{r}"].number_format = '0.0%'

    # ---------------------------------------------------------
    # BARIS TOTAL
    # ---------------------------------------------------------
    tot_row = total_rows + 2
    if total_rows > 0:
        cell_label  = ws_calc.cell(row=tot_row, column=1, value="TOTAL")
        cell_rev    = ws_calc.cell(row=tot_row, column=3, value=f"=SUM(C2:C{tot_row-1})")
        cell_cost   = ws_calc.cell(row=tot_row, column=4, value=f"=SUM(D2:D{tot_row-1})")
        cell_tax    = ws_calc.cell(row=tot_row, column=5, value=f"=SUM(E2:E{tot_row-1})")
        cell_profit = ws_calc.cell(row=tot_row, column=6, value=f"=SUM(F2:F{tot_row-1})")
        cell_margin = ws_calc.cell(row=tot_row, column=7, value=f"=AVERAGE(G2:G{tot_row-1})")

        # Apply Format Angka pada Baris Total
        cell_rev.number_format    = '#,##0'
        cell_cost.number_format   = '#,##0'
        cell_tax.number_format    = '#,##0'
        cell_profit.number_format = '#,##0'
        cell_margin.number_format = '0.0%'

        # Make Bold
        bold_font = Font(bold=True)
        cell_label.font  = bold_font
        cell_rev.font    = bold_font
        cell_cost.font   = bold_font
        cell_tax.font    = bold_font
        cell_profit.font = bold_font
        cell_margin.font = bold_font

    output = BytesIO()
    
    # AUTO-FIT LEBAR KOLOM (MENCEGAH TAMPILAN #####)
    for ws in [ws_master, ws_calc]:
        for col in ws.columns:
            max_len = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                # Memberi estimasi lebar jika sel berisi angka berformat/rumus
                if cell.number_format and cell.number_format != 'General':
                    max_len = max(max_len, len(val) + 8)
                else:
                    max_len = max(max_len, len(val))
            # Set lebar kolom dengan nilai minimal 15 agar rapi
            ws.column_dimensions[col_letter].width = max(max_len + 3, 15)
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Master_Kalkulasi_Rekap_Semua.xlsx"}
    )
